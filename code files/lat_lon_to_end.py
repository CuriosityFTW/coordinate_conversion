import numpy as np
import pymap3d as pm
# import pandas as pd
import openpyxl
import time

time_start = time.time()

wb = openpyxl.load_workbook("Single_point_data_testcase20_GPS_python.xlsx")
sheet = wb.sheetnames
ws = wb[sheet[0]]

pos_e = [0]*(ws.max_row - 1)
pos_n = [0]*(ws.max_row - 1)
pos_d = [0]*(ws.max_row - 1)

# Reference
for row in range(1, 2):
    for col in ws.iter_cols(0, 1):
        lat0 = col[row].value
    for col in ws.iter_cols(1, 2):
        lon0 = col[row].value    
    for col in ws.iter_cols(2, ws.max_column):
        h0 = col[row].value

# Target
for row in range(2, ws.max_row):
    for col in ws.iter_cols(0, 1):
        lat = col[row].value
    for col in ws.iter_cols(1, 2):
        lon = col[row].value    
    for col in ws.iter_cols(2, ws.max_column):
        h = col[row].value
    pm.geodetic2enu(lat, lon, h, lat0, lon0, h0)
    pos_e[row-2] = round(pm.geodetic2enu(lat, lon, h, lat0, lon0, h0)[0], 5)
    pos_n[row-2] = round(pm.geodetic2enu(lat, lon, h, lat0, lon0, h0)[1], 5)
    pos_d[row-2] = round(pm.geodetic2enu(lat, lon, h, lat0, lon0, h0)[2], 5)

pos_ne = [0]*len(pos_n)
pos_ne_square = [0]*len(pos_n)
pos_ne_sigma_temp = [0]*len(pos_n)
pos_ne_mean_sum = 0
pos_ne_square_sum = 0
pos_ne_sigma_sum = 0

for i in range(0, len(pos_n)):
    pos_ne[i] = round(np.sqrt(pos_n[i]**2 + pos_e[i]**2), 5)
    pos_ne_square[i] = pos_ne[i]**2
    pos_ne_mean_sum += pos_ne[i]
    pos_ne_square_sum += pos_ne_square[i]
    
pos_ne_mean = round(pos_ne_mean_sum / len(pos_n), 5)
pos_ne_square_mean = round(pos_ne_square_sum / len(pos_n), 5)

for i in range(0, len(pos_n)):
    pos_ne_sigma_temp[i] = (pos_ne[i] - pos_ne_mean)**2
    pos_ne_sigma_sum += pos_ne_sigma_temp[i]

pos_ne_sigma = round(np.sqrt(pos_ne_sigma_sum / (len(pos_n)-1)), 5)    
pos_ne_rms = round(np.sqrt(pos_ne_square_mean), 5)

# df = pd.DataFrame([[pos_e, pos_n, pos_d]], columns=['pos_e', 'pos_n', 'pos_d'])
# df.to_excel('lat_lon_to_end_testcase23_3.xlsx', sheet_name='lat_lon_to_end')

time_end = time.time()

print("Position mean_testcase21 =", pos_ne_mean, "m")
print("Position sigma_testcase21 =", pos_ne_sigma, "m")
print("Position rms_testcase21 =", pos_ne_rms, "m")
print("Computation time =", round((time_end - time_start)/60, 3), "min")